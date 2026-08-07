from __future__ import annotations

from collections import defaultdict
from datetime import UTC, date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook

from tca_web.application.contracts import JsonObject

MAX_ROWS = 100_000
MANIFEST_SHEET = "_manifest"


class WorkbookValidationError(ValueError):
    pass


def _value(value: object) -> str | int | float | bool | None:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def parse_workbook(content: bytes) -> list[JsonObject]:
    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as error:
        raise WorkbookValidationError("The file is not a valid .xlsx workbook.") from error
    records: list[JsonObject] = []
    try:
        for sheet in workbook.worksheets:
            if sheet.title == MANIFEST_SHEET:
                continue
            rows = sheet.iter_rows()
            header_cells = next(rows, ())
            headers = [str(cell.value or "").strip() for cell in header_cells]
            if not headers or not any(headers):
                continue
            if any(not header for header in headers) or len(headers) != len(set(headers)):
                raise WorkbookValidationError(
                    f"Sheet '{sheet.title}' has blank or duplicate column names."
                )
            for row_number, cells in enumerate(rows, 2):
                if any(cell.data_type == "f" for cell in cells):
                    raise WorkbookValidationError(
                        f"Sheet '{sheet.title}' row {row_number} contains a formula."
                    )
                values = [_value(cell.value) for cell in cells[: len(headers)]]
                if not any(value not in (None, "") for value in values):
                    continue
                records.append(
                    {
                        "_sheet": sheet.title,
                        **dict(zip(headers, values, strict=True)),
                    }
                )
                if len(records) > MAX_ROWS:
                    raise WorkbookValidationError(
                        f"Workbook exceeds the {MAX_ROWS:,}-row import limit."
                    )
    finally:
        workbook.close()
    return records


def export_workbook(
    records: list[JsonObject],
    *,
    workspace_id: str,
    company_id: int,
) -> bytes:
    workbook = Workbook()
    manifest = workbook.active
    if manifest is None:
        raise RuntimeError("Could not create workbook manifest")
    manifest.title = MANIFEST_SHEET
    manifest.append(["schema_version", 1])
    manifest.append(["workspace_id", workspace_id])
    manifest.append(["company_id", company_id])
    manifest.append(["exported_at_utc", datetime.now(UTC).isoformat()])
    grouped: defaultdict[str, list[JsonObject]] = defaultdict(list)
    for record in records:
        grouped[str(record.get("_sheet") or "History")].append(record)
    used_titles = {MANIFEST_SHEET.casefold()}
    for source_title, rows in grouped.items():
        title = _safe_title(source_title, used_titles)
        sheet = workbook.create_sheet(title)
        headers = sorted({key for row in rows for key in row if not key.startswith("_")})
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header) for header in headers])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _safe_title(source: str, used: set[str]) -> str:
    title = "".join("_" if character in r"[]:*?/\\" else character for character in source)
    title = title.strip()[:31] or "History"
    candidate = title
    suffix = 2
    while candidate.casefold() in used:
        marker = f"_{suffix}"
        candidate = f"{title[: 31 - len(marker)]}{marker}"
        suffix += 1
    used.add(candidate.casefold())
    return candidate
